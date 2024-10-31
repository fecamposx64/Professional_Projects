from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font



def atualizar_planilha():


    nome_coluna_nova = "Novembro 2024"
    
    wb_planilha = load_workbook('Hosts Monitorados.xlsx')
    ws_planilha = wb_planilha['Hosts Monitorados']

    wb_host = load_workbook('Hosts_Coletados.xlsx')
    ws_host = wb_host.active

    coluna_existente = False
    for col in ws_planilha.iter_rows(min_row=1, max_row=1, values_only=True):
        if nome_coluna_nova in col:
            coluna_existente = True
            print(f"A coluna '{nome_coluna_nova}' já existe. Nenhuma atualização realizada.")
            break

    if not coluna_existente:
        nova_coluna_index = ws_planilha.max_column + 1
        ws_planilha.cell(row=1, column=nova_coluna_index, value=nome_coluna_nova)

        # Aplica a cor de fundo e a cor da fonte
        fill_color = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        font_color = Font(color='000000', bold=True)

        # Formata o título da coluna
        title_cell = ws_planilha.cell(row=1, column=nova_coluna_index)
        title_cell.fill = fill_color
        title_cell.font = font_color

        # Preenche a nova coluna com os valores de df_host
        for index, row in enumerate(ws_host.iter_rows(min_row=2, values_only=True), start=2):
            host_value = row[0]  # Supondo que o valor do host está na primeira coluna
            ws_planilha.cell(row=index, column=nova_coluna_index, value=host_value)

        # Ajusta a largura da nova coluna
        max_length = max(len(nome_coluna_nova),  # Largura do título
                        *[len(str(row[0])) for row in ws_host.iter_rows(min_row=2, values_only=True)])  # Largura dos dados
        ws_planilha.column_dimensions[ws_planilha.cell(row=1, column=nova_coluna_index).column_letter].width = max_length + 2  # Ajusta a largura com um buffer extra


        wb_planilha.save('Hosts Monitorados.xlsx')


        print(f"Coluna '{nome_coluna_nova}' foi adicionada e preenchida com sucesso.")


atualizar_planilha()