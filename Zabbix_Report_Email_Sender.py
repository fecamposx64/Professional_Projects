import pandas as pd
import os
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
import win32com.client as client


def generate_report():
    """
    Gera um relatório de alertas de uso de disco, filtra os dados
    e exporta para um arquivo Excel.
    """
    print("Gerando o relatório de uso de disco...")

    # Carregar dados do arquivo CSV
    df = pd.read_csv("zbx_problems_export.csv")
    
    # Selecionar colunas relevantes
    relevant_columns = ['Time', 'Host', 'Problem', 'Severity']
    filtered_data = df[relevant_columns]

    # Filtrar alertas relacionados a "disk space"
    disk_alerts = filtered_data[filtered_data['Problem'].str.contains("disk space", case=False, na=False)]

    # Renomear colunas para facilitar a compreensão
    disk_alerts = disk_alerts.rename(columns={
        "Time": "Data/Hora", 
        "Host": "Hostname", 
        "Problem": "Alerta no Disco", 
        "Severity": "Severidade"
    })

    # Exibir o DataFrame filtrado
    print("Dados filtrados com sucesso:")
    print(disk_alerts)

    # Exportar os dados filtrados para um arquivo Excel
    disk_alerts.to_excel("Report Disk Usage Daily.xlsx", sheet_name='Sheet1', index=False)
    print("Relatório exportado para 'Report Disk Usage Daily.xlsx'.\n")


def report_formatting():
    """
    Formata o relatório gerado, aplicando estilos e ajustando a largura das colunas.
    """
    print("Aplicando formatação ao relatório...")

    # Carregar o arquivo Excel gerado para aplicar formatação
    wb = load_workbook("Report Disk Usage Daily.xlsx")
    ws = wb['Sheet1']

    # Definir o estilo do cabeçalho (cor de fundo e fonte)
    header_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Verde escuro
    header_font = Font(color="FFFFFF", bold=True)  # Fonte branca e em negrito

    # Aplicar o estilo ao cabeçalho
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Ajustar a largura das colunas automaticamente
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # Obtém a letra da coluna
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Adiciona espaço extra
        ws.column_dimensions[column_letter].width = adjusted_width

    # Salvar o arquivo Excel com a formatação aplicada
    wb.save("Report Disk Usage Daily.xlsx")
    print("Formatação aplicada com sucesso e arquivo salvo.\n")


def send_email_report():
    """
    Envia o relatório diário de uso de disco por e-mail aos clientes, com um relatório filtrado para cada cliente.
    """
    print("Iniciando envio de relatórios por e-mail...\n")
    
    try:
        # Construir o caminho absoluto do relatório
        report_file_path = os.path.join(Path.home(), "Desktop", "DailyDiskReport", "Report Disk Usage Daily.xlsx")
        
        # Carregar dados dos clientes e do relatório de disco
        df_cliente = pd.read_excel("contato_cliente.xlsx")
        df_report = pd.read_excel("Report Disk Usage Daily.xlsx")

        # Inicializar o Outlook
        outlook = client.Dispatch("Outlook.Application")
        msg_html = "Bom dia! Segue anexo o relatório diário de uso do disco."

        # Iterar sobre o DataFrame de clientes (df_cliente)
        for index, row in df_cliente.iterrows():
            cliente = row['Cliente']
            destinatario = row['Email']

            # Verificar se o cliente está presente na coluna 'Hostname' do relatório de disco (df_report)
            if df_report['Hostname'].str.contains(cliente, case=False, na=False).any():
                print(f"Cliente '{cliente}' encontrado. Preparando o relatório...")

                # Filtrar o relatório de disco para o cliente
                filtered_df_report = df_report[df_report['Hostname'].str.contains(cliente, case=False, na=False)]

                # Salvar o relatório filtrado para o cliente
                filtered_df_report.to_excel("Report Disk Usage Daily.xlsx", sheet_name='Sheet1', index=False)

                # Aplicar a formatação ao relatório
                report_formatting()
                
                # Criar e configurar o e-mail
                message = outlook.CreateItem(0)
                message.Recipients.Add(destinatario)
                message.CC = "<>"
                message.Subject = f"[{cliente}] Relatório diário de uso de disco"
                message.GetInspector
                index_body = message.HTMLbody.find('>', message.HTMLbody.find('<body'))
                message.HTMLbody = message.HTMLbody[:index_body + 1] + msg_html + message.HTMLbody[index_body + 1:]
                
                # Anexar o relatório ao e-mail
                message.Attachments.Add(report_file_path)

                # Exibir o e-mail antes de enviar (substituir por message.Send() para envio direto)
                message.Display()
                
                print(f"Relatório enviado para '{destinatario}'.\n")

            else:
                print(f"Cliente '{cliente}' não encontrado no relatório de disco.\n")
    
    except FileNotFoundError as e:
        print(f"Erro: Arquivo não encontrado - {e}")
    except Exception as e:
        print(f"Erro: Ocorreu um problema - {e}")


# Gerar o relatório e enviá-lo por e-mail
print("Iniciando processo de geração e envio de relatórios...\n")
generate_report()
send_email_report()
print("Processo concluído.")
